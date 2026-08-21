#!/usr/bin/env python3
"""
Build pl/selections.csv from the round-by-round prediction workbooks in
rounds/ROUND *.xlsm.

Each workbook has a header row (Match Number, FIXTURE, Actual Score, then one
column per player) followed by one row per fixture, e.g.:

    Match Number | FIXTURE            | Actual Score | Paul Knipe | Rhys Knipe | ...
    1            | Arsenal - Coventry |              | 3-0        | 2-2        | ...

The "Actual Score" column is ignored — pl/scores.csv is the source of truth
for results. The player list is the union across every round file (people
occasionally join/drop), in first-seen order; a player who didn't submit a
pick for a round gets a blank cell for that round's rows.

Output: pl/selections.csv, header `Round,Match No,FIXTURE,<player 1>,...`.

Usage:
  python3 tools/build_pl_selections.py [--rounds-dir rounds] [--output pl/selections.csv]
"""

from __future__ import annotations

import argparse
import csv
import re
import sys
from pathlib import Path

import openpyxl

HERE = Path(__file__).resolve().parent
ROOT = HERE.parent

ROUND_NUM_RE = re.compile(r"ROUND\s*(\d+)", re.IGNORECASE)


def round_number(path: Path) -> int:
    m = ROUND_NUM_RE.search(path.stem)
    if not m:
        raise ValueError(f"Can't find a round number in {path.name!r}")
    return int(m.group(1))


def find_header_row(ws) -> int:
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=1):
        cell = row[0]
        if isinstance(cell.value, str) and cell.value.strip() == "Match Number":
            return cell.row
    raise ValueError(f"No 'Match Number' header found in sheet {ws.title!r}")


def load_round(path: Path) -> tuple[int, list[str], list[dict[str, str]]]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.worksheets[0]
    header_row = find_header_row(ws)

    players: list[str] = []
    for col in range(4, ws.max_column + 1):
        value = ws.cell(row=header_row, column=col).value
        if isinstance(value, str) and value.strip():
            players.append(value.strip())

    rows: list[dict[str, str]] = []
    for r in range(header_row + 1, ws.max_row + 1):
        match_no = ws.cell(row=r, column=1).value
        if not isinstance(match_no, (int, float)):
            break
        fixture = ws.cell(row=r, column=2).value
        if not isinstance(fixture, str) or not fixture.strip():
            continue

        row = {"Match No": str(int(match_no)), "FIXTURE": fixture.strip()}
        for col, player in enumerate(players, start=4):
            value = ws.cell(row=header_row, column=col).value
            pick = ws.cell(row=r, column=col).value
            row[player.strip()] = str(pick).strip() if isinstance(pick, str) and pick.strip() else ""
        rows.append(row)

    return round_number(path), players, rows


def main() -> int:
    ap = argparse.ArgumentParser(description="Build pl/selections.csv from rounds/ROUND *.xlsm")
    ap.add_argument("--rounds-dir", type=Path, default=ROOT / "rounds")
    ap.add_argument("--output", type=Path, default=ROOT / "pl" / "selections.csv")
    args = ap.parse_args()

    paths = sorted(args.rounds_dir.glob("ROUND *.xlsm"), key=round_number)
    if not paths:
        print(f"Error: no ROUND *.xlsm files found in {args.rounds_dir}", file=sys.stderr)
        return 1

    all_players: list[str] = []
    all_rows: list[dict[str, str]] = []
    for path in paths:
        rnd, players, rows = load_round(path)
        for player in players:
            if player not in all_players:
                all_players.append(player)
        for row in rows:
            row["Round"] = str(rnd)
        all_rows.extend(rows)
        print(f"[i] {path.name}: round {rnd}, {len(rows)} fixtures, {len(players)} players")

    fieldnames = ["Round", "Match No", "FIXTURE", *all_players]
    args.output.parent.mkdir(parents=True, exist_ok=True)
    with args.output.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for row in all_rows:
            writer.writerow({name: row.get(name, "") for name in fieldnames})

    print(f"Wrote {args.output} ({len(all_rows)} rows, {len(all_players)} players)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
